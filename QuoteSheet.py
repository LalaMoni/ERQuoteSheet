import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from copy import copy
from PIL import Image
from io import BytesIO
import datetime
import pandas as pd
import uuid


# ---------------- 函数 ----------------
def write_cell_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
    else:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = merged_range.min_row, merged_range.min_col
                ws.cell(row=top_left[0], column=top_left[1], value=value)
                break

def calculate_prices(P, product_Q, total_Q, F, R):
    A_CNY = P + F / total_Q
    B_CNY = P / 1.13 * 1.02 + F / total_Q
    A_USD = A_CNY / R
    B_USD = B_CNY / R
    return round(B_CNY, 4), round(A_CNY, 4), round(B_USD, 4), round(A_USD, 4)

def excel_cell_size_to_pixels(ws, row, col):
    col_letter = ws.cell(row=row, column=col).column_letter
    col_width = ws.column_dimensions[col_letter].width or 8.43
    row_height = ws.row_dimensions[row].height or 15
    px_width = int(col_width * 7)
    px_height = int(row_height * 1.33)
    return px_width, px_height

def insert_image(ws, img_file, cell, max_width, max_height, scale=0.7):
    img_pil = Image.open(BytesIO(img_file.getbuffer()))
    img_pil.thumbnail((max_width * scale, max_height * scale))
    img_bytes = BytesIO()
    img_pil.save(img_bytes, format="PNG")
    img_bytes.seek(0)
    img = XLImage(img_bytes)
    ws.add_image(img, cell)

# ---------------- Streamlit 页面 ----------------
st.title("报价单生成器")

st.header("上传 Excel 模板")
uploaded_template = st.file_uploader("请选择 Excel 模板文件", type=["xlsx"])

# 基本信息
st.header("基本信息")
purchaser = st.text_area("采购商信息")
today = datetime.date.today()
order_no = st.text_input("编号", value=f"ERKJ{today.strftime('%Y%m%d')}XX")
date_input = st.text_input("日期", value=today.strftime("%Y/%m/%d"))
F_input = st.text_input("总费用（支持公式，例如 200+50*4）")
R_input = st.text_input("汇率（买入价）")
start_row = 13  # 默认13行

# --- 产品信息 ---
st.header("产品信息")
product_options = {
    "吸气片": ["SG-01", "SG-02", "SG-03"],
    "焊料": ["CB-01", "CB-02"],
}

# 初始化
if "products" not in st.session_state:
    st.session_state.products = [{
        "uid": str(uuid.uuid4()),
        "name": list(product_options.keys())[0],
        "model": product_options[list(product_options.keys())[0]][0],
        "P": None,
        "Q": None,
        "img": None
    }]
    
if "product_images" not in st.session_state:
    st.session_state.product_images = [p["img"] for p in st.session_state.products]

products = st.session_state.products

for i, p in enumerate(st.session_state.products):
    st.markdown("---")
    col_title, col_up, col_down, col_del = st.columns([6, 1, 1, 1])
    with col_title:
        st.subheader(f"产品 {i+1}")  
    with col_up:
        if st.button("上移", key=f"up{i}", disabled=(i == 0)):
            products[i - 1], products[i] = products[i], products[i - 1]
            st.session_state.product_images[i - 1], st.session_state.product_images[i] = \
                st.session_state.product_images[i], st.session_state.product_images[i - 1]
            st.rerun()
    with col_down:
        if st.button("下移", key=f"down{i}", disabled=(i == len(products) - 1)):
            products[i + 1], products[i] = products[i], products[i + 1]
            st.session_state.product_images[i + 1], st.session_state.product_images[i] = \
                st.session_state.product_images[i], st.session_state.product_images[i + 1]
            st.rerun()
    with col_del:
        if st.button("删除", key=f"del{i}"):
            del st.session_state.products[i]
            del st.session_state.product_images[i]
            st.rerun()
                
    # 输入产品信息
    name = st.selectbox("产品名称", list(product_options.keys()),
                        index=list(product_options.keys()).index(p["name"]),
                        key=f"name_{p['uid']}")
    model = st.selectbox("型号", product_options[name],
                         index=product_options[name].index(p["model"]) if p["model"] in product_options[name] else 0,
                         key=f"model_{p['uid']}")
    P = st.number_input("净单价", value=p["P"], format="%.4f", key=f"P_{p['uid']}")
    Q = st.number_input("数量", value=p["Q"] if p["Q"] is not None else 0, min_value=0, step=1, key=f"Q_{p['uid']}")
    uploaded_file = st.file_uploader("上传图片", type=["png","jpg","jpeg"], key=f"img_{p['uid']}")

    if uploaded_file is not None:
        st.session_state.product_images[i] = uploaded_file
    p.update({"name": name, "model": model, "P": P, "Q": Q, "img": st.session_state.product_images[i]})

    # 添加产品按钮
    if st.button(f"添加产品", key=f"add_after_{i}"):
        new_product = {
            "uid": str(uuid.uuid4()),
            "name": list(product_options.keys())[0],
            "model": product_options[list(product_options.keys())[0]][0],
            "P": None,
            "Q": None,
            "img": None
        }
        products.insert(i+1, new_product)
        st.session_state.product_images.insert(i+1, None)
        st.rerun()


# ---------------- 预览&生成 ----------------
col_space, col_buttons = st.columns([6, 3])

with col_buttons:
    col_preview, col_generate = st.columns([1, 1])

    # 预览报价单
    with col_preview:
        if st.button("预览报价单"):
            try:
                F = eval(F_input)
            except Exception as e:
                st.error(f"总费用输入错误: {e}")
                st.stop()
            try:
                R = float(R_input)
            except:
                st.error("汇率输入错误")
                st.stop()

            total_Q = sum(p["Q"] for p in products if p["Q"])
            preview_data = []
            for p in products:
                B_CNY, A_CNY, B_USD, A_USD = calculate_prices(p["P"], p["Q"], total_Q, F, R)
                preview_data.append({
                    "产品": p["name"],
                    "型号": p["model"],
                    "数量": p["Q"],
                    "人民币单价(不含税)": B_CNY,
                    "人民币单价(含税)": A_CNY,
                    "美元单价(不含税)": B_USD,
                    "美元单价(含税)": A_USD
                })
            df_preview = pd.DataFrame(preview_data)
            st.table(df_preview)

    # 生成报价单
    with col_generate:
        if st.button("生成报价单"):
            if not uploaded_template:
                st.error("请先上传 Excel 模板")
                st.stop()
            
            try:
                F = eval(F_input)
                R = float(R_input)
            except Exception as e:
                st.error(f"F 或 R 输入错误: {e}")
                st.stop()
    
            total_Q = sum(p["Q"] for p in products)
            wb = load_workbook(BytesIO(uploaded_template.read()))
            ws = wb.active
        
            # 写入基本信息
            write_cell_safe(ws, 4, 7, purchaser)
            ws.cell(row=8, column=7, value=order_no)
            ws.cell(row=9, column=7, value=date_input)

            # 列位置
            NO_COL, PRODUCT_COL, IMG_COL, MODEL_COL, QUANTITY_COL = 1, 2, 3, 4, 5
            RMB_COL_START, USD_COL_START = 7, 9
    
            # 写入产品数据
            for p in products:
                row = start_row + idx
                ws.row_dimensions[row].height = 69
                max_w, max_h = excel_cell_size_to_pixels(ws, row, IMG_COL)
    
                ws.cell(row=row, column=NO_COL, value=idx + 1)
                ws.cell(row=row, column=PRODUCT_COL, value=p["name"])
                ws.cell(row=row, column=MODEL_COL, value=p["model"])
                ws.cell(row=row, column=QUANTITY_COL, value=p["Q"])

                B_CNY, A_CNY, B_USD, A_USD = calculate_prices(p["P"], p["Q"], total_Q, F, R)
                if B_CNY:
                    write_cell_safe(ws, row, RMB_COL_START, B_CNY)
                    write_cell_safe(ws, row, RMB_COL_START + 1, A_CNY)
                    write_cell_safe(ws, row, USD_COL_START, B_USD)
                    write_cell_safe(ws, row, USD_COL_START + 1, A_USD)
    
                # 插入图片
                if p["img"]:
                    insert_image(ws, p["img"], f"C{row}", max_width=max_w, max_height=max_h)
                
            # 保存新文件到 BytesIO 提供下载
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            new_file_name = f"报价单_{date_input.replace('/', '-')}.xlsx"
    
            st.download_button(
                "下载报价单",
                data=output,
                file_name=new_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
